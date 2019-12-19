#! python3

# Made by Seth Hannah
API_KEY = 'ENTER GOOGLE API KEY HERE'
try:
    from PIL import Image
except:
    import Image
import os
import re
import smtplib
import wx
import cloudvisreq as cvr
import json
import openpyxl

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

class Utils():
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def getImagePaths():
        ret = []
        dirs = os.listdir(os.getcwd())
        for dir in dirs:
            if ".jpg" in dir or ".jpeg" in dir:
                ret.append(dir)
        return ret

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def testImages(image_filenames):

        # get images json info from cloud api
        response = cvr.request_ocr(API_KEY, image_filenames)
        print (image_filenames)

        # if request fails raise exception
        if response.status_code != 200 or response.json().get('error'):
            raise Exception(response.text)

        # if request succeeds build array
        else:
            ret = []

            # loop through all the responses
            for idx, resp in enumerate(response.json()['responses']):
                try:
                    imgstr = resp['textAnnotations'][0]['description']
                    print (imgstr, "\n\n")
                except:
                    # if no text is found append error so picture can be removed
                    ret.append("error")
                    continue

                imgreg1 = re.search(r"(M{0,1}[0-9]{3,4}[A-Z]{0,2}[-]{1}[0-9A-Z]{1,4})", imgstr)
                imgreg2 = re.search(r"(M[0-9]{3,4})", imgstr)
                if imgreg1 != None or imgreg2 != None:

                #add something in to be able to find codes that look like MXXXX etc.

                    if (imgreg1 != None):
                        ret.append(imgreg1.group(1))

                    elif(imgreg2 != None):
                        ret.append(imgreg2.group(1))

                    else:
                        ret.append("")
                else:
                    ret.append("")

        return ret



    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def prepValues():
        dirs = Utils.getImagePaths()
        vals = []
        print (len(dirs))
        for x in range(0, len(dirs), 10):
            vals.extend(Utils.testImages(dirs[x:x+10]))
            print (vals)
        print ("Total Vals:\n", vals)
        ret = {}
        for x in range(len(dirs)):
            print (x)
            if vals[x] != "error":
                ret[dirs[x]] = vals[x]

        return ret

    def diff(first, second):
        second = set(second)
        return [item for item in first if item not in second]

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

class MainWindow(wx.Frame):
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def __init__(self, parent, title):
        wx.Frame.__init__(self, parent, title=title, size=(200, 100))
        self.initUI()

        self.picset = self.initCodes()
        self.initForms()

        #self.emailpanel = EmailPanel(self)
        #self.emailpanel.Hide()
        #self.panel.Layout()
        self.panel.Show()
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def initUI(self):

        self.panel = wx.Panel(self)

        font = wx.SystemSettings.GetFont(wx.SYS_SYSTEM_FONT)
        font.SetPointSize(12)

        vbox1 = wx.BoxSizer(wx.VERTICAL)

        # ~~ BOX/ROW ONE ~~
        hbox1 = wx.BoxSizer(wx.HORIZONTAL)

        self.codebox = wx.TextCtrl(self.panel, style=wx.TE_PROCESS_ENTER)
        self.codebox.SetFont(font)
        hbox1.Add(self.codebox, flag=wx.EXPAND|wx.LEFT, border=5, proportion=1)

        self.prevbtn = wx.Button(self.panel, label="Prev")
        self.prevbtn.SetFont(font)
        hbox1.Add(self.prevbtn, flag=wx.RIGHT|wx.CENTER, border=5)

        self.nextbtn = wx.Button(self.panel, label="Next")
        self.nextbtn.SetFont(font)
        hbox1.Add(self.nextbtn, flag=wx.RIGHT|wx.CENTER, border=5)

        vbox1.Add(hbox1, flag=wx.EXPAND|wx.LEFT|wx.RIGHT|wx.TOP, border=10)

        vbox1.Add((-1, 5)) # ADD SPACER

        # ~~ BOX/ROW TWO ~~
        hbox2 = wx.BoxSizer(wx.HORIZONTAL)

        self.imagecontainer = wx.StaticBitmap(self.panel, size = wx.Size(500, 400))
        hbox2.Add(self.imagecontainer, flag=wx.CENTER|wx.EXPAND, border=50)

        vbox1.Add(hbox2, proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND, border=0)

        vbox1.Add((-1, 10)) # ADD SPACER

        # ~~ BOX/ROW THREE ~~
        hbox3 = wx.BoxSizer(wx.HORIZONTAL)

        self.counter = wx.StaticText(self.panel, label="X/X")
        self.counter.SetFont(font)
        hbox3.Add(self.counter, proportion=1, flag=wx.RIGHT)

        vbox1.Add(hbox3, proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND)

        self.panel.SetSizer(vbox1)

        self.nextbtn.Bind(wx.EVT_BUTTON, self.onNext)
        self.prevbtn.Bind(wx.EVT_BUTTON, self.onPrev)

        self.Bind(wx.EVT_TEXT_ENTER, self.onEnter, self.codebox)
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def setPicture(self, path):
        # set the base width of the picture for it to be scaled down to
        basewidth = 400
        img = Image.open(path)
        wpercent = (basewidth / float(img.size[0]))
        hsize = int((float(img.size[1]) * float(wpercent)))
        img = img.resize((basewidth, hsize), Image.ANTIALIAS)
        # save resized picture as resized_[path]
        img.save('resized_' + path)
        self.imagecontainer.SetBitmap(wx.Image("resized_" + path, wx.BITMAP_TYPE_JPEG).ConvertToBitmap())

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def initForms(self):
        self.codebox.SetValue(self.picset.getCurrentPicture().getCode())
        self.setPicture(self.picset.getCurrentPicture().getPath())
        self.counter.SetLabel("{}/{}".format(self.picset.getCurrentPictureIndex()+1, self.picset.totalPictures()))

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def initCodes(self):
        picarr = []
        for key, value in Utils.prepValues().items():
            picarr.append(Picture(key, value))
        return PictureSet(picarr)

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def onEnter(self, event):
        self.onNext(event)

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def onNext(self, event):
        if self.codebox.GetValue() != "":
            self.changeCode(1)
        else:
            print("cant have blank code")

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def onPrev(self, event):
        if self.codebox.GetValue() != "":
            self.changeCode(-1)
        else:
            print("cant have blank code")

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def changeCode(self, value):
        flag = True
        self.picset.getCurrentPicture().setCode(self.codebox.GetValue())
        if self.picset.changePicture(value):

            self.codebox.SetValue(self.picset.getCurrentPicture().getCode())

            self.setPicture(self.picset.getCurrentPicture().getPath())

            self.codebox.SetFocus()

            self.counter.SetLabel("{}/{}".format(self.picset.getCurrentPictureIndex()+1, self.picset.totalPictures()))

        else:
            if not self.picset.getCurrentPictureIndex() < 1 and flag:
                flag = False
                ret = []
                for pic in self.picset.pictures:
                    code = re.sub('[-]', '', pic.getCode())
                    ret.append(code)

                # detele resized picture
                os.system(" ".join(['rm', "resize*"]))
                #self.showEmailPanel()
                #Utils.sendEmail(ret)
                self.editExcelSheet(ret)

            else:
                print("go next instead of prev")

    def showEmailPanel(self):
        self.panel.Hide()
        self.emailpanel.Show()
        self.panel.Layout()
        self.emailpanel.Layout()

    def editExcelSheet(self, codes):
        wb = openpyxl.load_workbook('/home/seth/Development/python/gui/inputsheet.xlsx')
        count = 0
        foundcodes = []
        sheets = wb.sheetnames

        #2x3 laminate
        sheet = wb[sheets[0]]
        for row in range(2, sheet.max_row):
            for code in codes:
                if sheet['B' + str(row)].value:
                    if code in sheet['B' + str(row)].value:
                        print ("found " + code + "!")
                        foundcodes.append(code)
                        sheet['C' + str(row)] = 10
                        count += 1

        #2x3 decometal
        sheet = wb[sheets[7]]
        for row in range(2, sheet.max_row):
            for code in codes:
                if sheet['B' + str(row)].value:
                    if code in sheet['B' + str(row)].value:
                        print ("found " + code + "!")
                        foundcodes.append(code)
                        sheet['C' + str(row)] = 10
                        count += 1


        if count == len(codes):
            print ("all codes found")
        else:
            codesleftover = Utils.diff(codes, foundcodes)
            print("Codes leftover:")
            print(codesleftover)
        company = os.path.relpath(".", "..").title()
        company = company[:company.index('-')]
        wb.save('/home/seth/' + company.strip(' ') + '.xlsx')



#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

class EmailPanel(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent)
        self.initUI()
        self.parent = parent

    def initUI(self):

        font = wx.SystemSettings.GetFont(wx.SYS_SYSTEM_FONT)
        font.SetPointSize(12)

        vbox1 = wx.BoxSizer(wx.VERTICAL)

        # ~~ BOX/ROW ONE ~~
        hbox1 = wx.BoxSizer(wx.HORIZONTAL)

        self.email = wx.TextCtrl(self, style=wx.TE_PROCESS_ENTER)
        self.email.SetFont(font)
        hbox1.Add(self.email, flag=wx.EXPAND|wx.LEFT, border=5, proportion=1)

        self.send = wx.Button(self, label="Send")
        self.send.SetFont(font)
        hbox1.Add(self.send, flag=wx.RIGHT|wx.CENTER, border=5)

        vbox1.Add(hbox1, flag=wx.EXPAND|wx.LEFT|wx.RIGHT|wx.TOP, border=10)

        vbox1.Add((-1, 5)) # ADD SPACER

        # ~~ BOX/ROW TWO ~~
        hbox2 = wx.BoxSizer(wx.HORIZONTAL)

        vbox1.Add(hbox2, proportion=1, flag=wx.LEFT|wx.RIGHT|wx.EXPAND, border=0)

        vbox1.Add((-1, 10)) # ADD SPACER



class PictureSet(object):
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def __init__(self, pictures=[]):
        self.pictures = pictures
        self.currentindex = 0
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def totalPictures(self):
        return len(self.pictures)
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def next(self):
        #add 1 to picture count, return true if successful or false if unsuccessful
        if self.currentindex < self.totalPictures():
            self.currentindex += 1
            return True
        else:
            return False

    def changePicture(self, value):
        #add 1 to picture count, return true if successful or false if unsuccessful
        flag = self.currentindex + value
        if flag >= 0 and flag < self.totalPictures():
            self.currentindex += value
            return True
        else:
            return False

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def getCurrentPicture(self):
        return self.pictures[self.currentindex]

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def getCurrentPictureIndex(self):
        return self.currentindex

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

class Picture(object):
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def __init__(self, path, code=""):
        self.path = path
        self.code = code

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def getCode(self):
        return self.code

    def getPath(self):
        return self.path

    def setCode(self, code):
        self.code = code

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------



def main():

    app = wx.App()
    ex = MainWindow(None, "Small editor")
    ex.Show()
    app.MainLoop()


if __name__ == '__main__':
    main()
